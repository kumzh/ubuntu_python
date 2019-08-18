import openpyxl
import os
from sys import argv
from openpyxl.utils import get_column_letter
#加载文件
def load_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    flag = True 
    while flag:
        # sheet_name = input('input %s sheet name:'%file_name)
        col = int(input('输入要匹配的列 编号：'))
        try:
            # sheet = wb.get_sheet_by_name(sheet_name)
            sheet = wb.get_active_sheet()
            flag = False
        except Exception as e:
            print (e)
    return sheet,col,wb
    

def file_test(file1,file2):
    if os.path.exists(file1) and os.path.exists(file2):
        pass
    else:
        print('文件不在当前文件夹')
        exit(1)

def pull_data(sheet,col):
    file_dict = {}
    for row_num in range(2,sheet.max_row+1):
        # file1_dict.setdefault(sheet1.cell(row = row_num,column = 1).value,{})
        file_dict[sheet.cell(row = row_num,column = 1).value] = sheet.cell(row = row_num,column = col).value
    return file_dict
        

def modify_file(file1,file2):
    print("load file1")
    sheet1 ,col1,wb1= load_file(file1)
    print("load file2")
    sheet2 ,col2,wb2= load_file(file2)

    file1_dict = pull_data(sheet1,col1)
    file2_dict = pull_data(sheet2,col2)
    
    print("before ",file1_dict,file2_dict)
    for id in file1_dict.keys():
        for id2 in file2_dict.keys():
            if id == id2:
                file2_dict[id2] = file1_dict[id]

    for row_num in range(2,sheet2.max_row+1):
        a = get_column_letter(col2)
        con = a + str(row_num)
        sheet2[con] = file2_dict[sheet2.cell(row = row_num,column = 1).value]
    wb2.save(file2)
    # print("late ",file1_dict,file2_dict)

    
    
    



if __name__ == "__main__":
    if len(argv) != 3:
        print("参数错误")
        exit(1)
    else:
        script_name,file1,file2 = argv
        file_test(file1,file2)
        print(script_name + 'is running...'+'\n' + '*'*20)
        modify_file(file1,file2)
        


