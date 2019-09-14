import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import fonts,stylesheet

def load_file(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    # fontobj = fonts(name="Times New Roman",bold=True)
    for i in range(1,10):
        con1 = "A" + str(i+1)
        # sheet[con1].style/styleobj
        sheet[con1] = i
        con2 = get_column_letter(i+1) + "1"
        # sheet[con2].style/styleobj
        sheet[con2] = i
    for i in range(1,10):
        a = sheet.cell(row = i+1,column = 1).value
        for j in range(1,10):
            b = sheet.cell(row = 1,column = j+1).value
            c = a * b
            con3 = get_column_letter(j+1) + str(i+1)
            sheet[con3] = c

    wb.save(file)
if __name__ == "__main__":
    load_file("9*9.xlsx")