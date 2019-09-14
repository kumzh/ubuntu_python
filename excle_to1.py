import openpyxl
import os
from sys import argv
from openpyxl.utils import get_column_letter

def open_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet_list = wb.get_sheet_names()
    print(sheet_list,len(sheet_list))
    return wb,sheet_list


if __name__ == "__main__":
    a = len(argv)
    print(a)
    if a == 1:
        print('请加上要合并的文件！')
        exit(1)
    else:
        file_create = openpyxl.Workbook()
        write_sheet = file_create.get_active_sheet()
        row_count = 0
        print()
        for i in range(1,a):
            wb,sheet_list = open_file(argv[i])
            for sheet_name in sheet_list:
                sheet = wb.get_sheet_by_name(sheet_name)
                a = sheet.max_row
                b = sheet.max_column
                for row_num in range(1,a+1):
                    row_count += 1
                    for col_num in range(1,b+1):
                        con = get_column_letter(col_num)+str(row_count)
                        write_sheet[con] = sheet.cell(row=row_num,column=col_num).value

            
        print(a-1)
        file_create.save('new.xlsx')